import streamlit as st
import pandas as pd
import re
from io import BytesIO
import time
from openpyxl import load_workbook

# Constants
ASSET_COLUMN = 'รหัสทรัพย์สิน'
CENTRAL_ASSET_COLUMN = 'หน่วยงานกลางรับทราบและตรวจสอบ'
FILTER_VALUES = ['อภิสรา สีดาคุณ']
INCORRECT_TERMS = {
    "รหัสทรัพย์สิน", "ไม่มี", "Computer", "Notebook", "ไม่มีรหัสทรัพย์สิน",
    "แทบเล็ต", "รายละเอียด", "nan", "-", "์Notebook", "ทบ. 5589338", "ทบ. 5589339 (รูปเครื่อง)"
}

# --- Email summary function ---
def summarize_email_domains(df, column_name='E-mail ผู้สร้างเอกสาร'):
    def extract_domain(email):
        if pd.isna(email) or str(email).strip() == "":
            return "cpall.co.th"
        match = re.search(r'@([\w\.-]+)', str(email))
        return match.group(1).lower() if match else "unknown"

    if column_name not in df.columns:
        return pd.DataFrame(columns=["Domain", "Count"])

    domain_counts = df[column_name].apply(extract_domain).value_counts().reset_index()
    domain_counts.columns = ["Domain", "Count"]
    return domain_counts

# --- Fallback Excel reader ---
def safe_read_excel_sheets(path):
    try:
        return pd.read_excel(path, sheet_name=None)
    except Exception as e:
        print(f"[WARN] Standard read failed: {e}")
        print("[INFO] Trying to recover using openpyxl read-only mode...")
        try:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            dataframes = {}
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                df = pd.DataFrame(rows[1:], columns=rows[0])
                dataframes[sheet] = df
            print("[INFO] Recovery succeeded.")
            return dataframes
        except Exception as ex:
            print(f"[ERROR] Recovery failed: {ex}")
            raise

# --- Filter helpers ---
def is_invalid_asset(asset_value):
    if pd.isna(asset_value) or not isinstance(asset_value, str) or asset_value.strip() == '':
        return True
    if asset_value.isdigit() and int(asset_value) == 0:
        return True
    if not asset_value.isdigit():
        return True
    if any(term in asset_value for term in INCORRECT_TERMS):
        return True
    return False

def cleanse_asset_code(asset_value):
    return re.sub(r'\D', '', asset_value) if isinstance(asset_value, str) else asset_value

def ensure_utf8_encoding(df):
    return df.apply(lambda x: x.str.encode('utf-8').str.decode('utf-8') if x.dtype == "object" else x)

# --- Main processing function ---
def process_excel(df):
    progress_bar = st.progress(0)

    incorrect_mask = df[ASSET_COLUMN].apply(is_invalid_asset)
    correct_data = df[~incorrect_mask].copy()
    incorrect_data = df[incorrect_mask].copy()

    correct_data[ASSET_COLUMN] = correct_data[ASSET_COLUMN].apply(cleanse_asset_code)
    progress_bar.progress(40)

    split_rows = []
    for _, row in incorrect_data.iterrows():
        assets = re.split(r'[ ,/\\*]+', str(row[ASSET_COLUMN]))
        for asset in filter(None, map(str.strip, assets)):
            cleaned_asset = cleanse_asset_code(asset)
            if cleaned_asset:
                new_row = row.copy()
                new_row[ASSET_COLUMN] = cleaned_asset
                split_rows.append(new_row)

    split_result = pd.DataFrame(split_rows)
    progress_bar.progress(60)

    merged_result = pd.concat([correct_data, split_result], ignore_index=True)
    merged_result["Duplicate"] = merged_result.duplicated(subset=[ASSET_COLUMN], keep=False).map({True: "Yes", False: "No"})

    tara_silom_data = merged_result[merged_result[CENTRAL_ASSET_COLUMN] == 'อภิสรา สีดาคุณ']
    duplicate_wrong_data = merged_result[
        (merged_result["Duplicate"] == "Yes") | (merged_result[ASSET_COLUMN].apply(is_invalid_asset))
    ]
    correct_all = merged_result.copy()

    for df in [correct_all, tara_silom_data, duplicate_wrong_data]:
        df[:] = ensure_utf8_encoding(df)

    progress_bar.progress(80)

    # === Generate Email Domain Summary AFTER all filtering ===
    email_summary = summarize_email_domains(correct_all, column_name='E-mail ผู้สร้างเอกสาร')

    # === Write all outputs ===
    output_buffer = BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        correct_all.to_excel(writer, sheet_name="Correct Data", index=False)
        tara_silom_data.to_excel(writer, sheet_name="Tara-Silom", index=False)
        duplicate_wrong_data.to_excel(writer, sheet_name="Duplicate & Wrong Data", index=False)
        email_summary.to_excel(writer, sheet_name="Company Email", index=False)

    output_buffer.seek(0)
    progress_bar.progress(100)
    time.sleep(0.5)
    progress_bar.empty()
    return output_buffer

# --- Streamlit UI ---
st.title("ALLProcess Data Cleaner V3.0")
uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

sheet_names = []
sheets = {}
if uploaded_file:
    try:
        sheets = safe_read_excel_sheets(uploaded_file)
        sheet_names = list(sheets.keys())
    except Exception as e:
        st.error(f"Failed to load Excel file: {e}")

if sheet_names:
    sheet_name = st.selectbox("Select the sheet to filter", sheet_names)
else:
    sheet_name = None

if uploaded_file and sheet_name:
    st.success(f"Processing sheet: {sheet_name}")
    df = sheets[sheet_name]
    output_file = process_excel(df)

    st.download_button(
        label="Download Processed File",
        data=output_file,
        file_name="cleaned_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
